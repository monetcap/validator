import logging
from openpyxl import load_workbook
from openpyxl import Workbook as PyxlWorkbook
import argparse
import re

PHONE_COLUMN_REGEX = r'^phone$'
PHONE_DIGIT_REGEX = r'^\d{10}$'

PHONE_REGEXS = [
    PHONE_COLUMN_REGEX,
    PHONE_DIGIT_REGEX
]

LOG_FORMAT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
LOG_LEVEL = 'DEBUG'

# configure logging
logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
log = logging.getLogger("validator")

parser = argparse.ArgumentParser(description='Process an xlsx document.')

parser.add_argument('--xlsx-file-path', '-xfp',
                    help='Path to xlsx document',
                    required=True
                    )


class Sheet:
    phone_column_index = -1

    valid_rows = []
    invalid_rows = []

    def _get_phone_column_index_from_row(self, row):
        # initial phone_column_index value
        phone_column_index = -1

        # generate cell values from row
        cell_values = self._get_cell_values_from_row(row)

        # iterate through cell values
        for i in range(len(cell_values)):
            value = cell_values[i]

            # Check if value matches "[Pp]hone || 0000000000"
            if self._is_value_phone(value):
                phone_column_index = i
                break

        return phone_column_index

    def _is_value_phone(self, value):
        return any(re.match(regex, value, re.IGNORECASE)
                   for regex in PHONE_REGEXS)

    def _get_cell_values_from_row(self, row):
        return [str(cell.value) for cell in row]

    def _enumerate_phone_column(self):
        # get first two rows from sheet
        rows = self.sheet[1:2]

        # this will loop through the first two rows and set the
        # phone_column_index when it's first enumerated
        for row in rows:
            value = self._get_phone_column_index_from_row(row)

            if value is not -1 and self.phone_column_index is -1:
                self.phone_column_index = value
                break

        # raise an exception if the value wasn't enumerated
        if self.phone_column_index is -1:
            raise Exception("phone_column_index couldn't be enumerated")

    def __init__(self, sheet):
        self.title = sheet.title
        self.sheet = sheet

        self._enumerate_phone_column()

        # log to debug initialization of Sheet object
        log.debug('<Sheet title={0} phone_column_index={1}>'.format(
            self.title, self.phone_column_index
        ))

    def validate(self):
        for row in self.sheet.rows:
            self._validate_row(row)

        self._write_results_to_workbook(self.valid_rows, 'valid')
        self._write_results_to_workbook(self.invalid_rows, 'invalid')

        log.debug('invalid_rows:\n{}'.format(self.invalid_rows))
        log.debug('valid_rows:\n{}'.format(self.valid_rows))

    def _validate_row(self, row):
        cell_values = self._get_cell_values_from_row(row)

        phone_value = cell_values[self.phone_column_index]

        if not re.match(PHONE_DIGIT_REGEX, phone_value):
            return

        if self._dnc_validate_phone(phone_value):
            self.valid_rows.append(cell_values)
        else:
            self.invalid_rows.append(cell_values)

    def _dnc_validate_phone(self, phone):
        if phone == '5051010101':
            return True
        else:
            return False

    def _write_results_to_workbook(self, results, workbook_name):
        wb = PyxlWorkbook()

        ws = wb.active
        ws.title = workbook_name

        for res in results:
            ws.append(res)

        wb.save(filename="/tmp/out/" + workbook_name + ".xlsx")


class Workbook:
    def __init__(self, file_path):
        self.workbook = load_workbook(file_path, read_only=True)
        self.sheets = [Sheet(sheet) for sheet in self.workbook]

    def validate(self):
        for sheet in self.sheets:
            sheet.validate()


class Validator:
    sheets = []

    def __init__(self, xlsx_file_path):
        self.workbook = Workbook(xlsx_file_path)

    def validate(self):
        self.workbook.validate()


def main():
    args = parser.parse_args()

    validator = Validator(xlsx_file_path=args.xlsx_file_path)
    validator.validate()


if __name__ == '__main__':
    main()
